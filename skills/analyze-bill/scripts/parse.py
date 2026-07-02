"""账单解析模块。

支持支付宝账单（CSV 格式）和微信账单（Excel 格式）的自动识别与解析，
将不同来源的账单统一转换为标准化的字段结构。
"""

import csv
import json
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

# 统一字段名称到中文显示名的映射。
# 键为内部标准化字段名，值为对应的中文标签。
FIELDS = {
    "time": "交易时间",
    "type": "交易类型",
    "target": "交易对方",
    "target_account": "对方账号",
    "description": "商品说明",
    "income_expense": "收/支",
    "amount": "金额",
    "payment_method": "支付方式",
    "status": "交易状态",
    "order_id": "交易订单号",
    "merchant_order_id": "商家订单号",
    "remark": "备注",
    "source": "来源",
}

# 支付宝账单原始字段名到内部标准化字段名的映射。
ALIPAY_MAPPING = {
    "交易时间": "time",
    "交易分类": "type",
    "交易对方": "target",
    "对方账号": "target_account",
    "商品说明": "description",
    "收/支": "income_expense",
    "金额": "amount",
    "收/付款方式": "payment_method",
    "交易状态": "status",
    "交易订单号": "order_id",
    "商家订单号": "merchant_order_id",
    "备注": "remark",
}

# 微信账单原始字段名到内部标准化字段名的映射。
WECHAT_MAPPING = {
    "交易时间": "time",
    "交易类型": "type",
    "交易对方": "target",
    "商品": "description",
    "收/支": "income_expense",
    "金额(元)": "amount",
    "支付方式": "payment_method",
    "当前状态": "status",
    "交易单号": "order_id",
    "商户单号": "merchant_order_id",
    "备注": "remark",
}

def is_alipay_bill(file_path: Path) -> bool:
    """判断文件是否为支付宝账单。

    通过读取文件头部内容，检测是否包含支付宝账单的特征关键词。

    Args:
        file_path: 待检测的文件路径。

    Returns:
        如果文件被识别为支付宝账单则返回 True，否则返回 False。
    """
    try:
        with open(file_path, "r", encoding="gbk") as f:
            text = f.read(4096)
        keywords = [
            "支付宝账户",
            "电子客户回单",
            "交易分类",
            "交易订单号",
        ]
        return sum(k in text for k in keywords) >= 3
    except Exception:
        return False

def is_wechat_bill(file_path: Path) -> bool:
    """判断文件是否为微信账单。

    通过 openpyxl 读取 Excel 文件的前 30 行，检测是否包含微信支付账单的特征关键词。

    Args:
        file_path: 待检测的 Excel 文件路径。

    Returns:
        如果文件被识别为微信账单则返回 True，否则返回 False。
    """
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        values = []
        for row in sheet.iter_rows(max_row=30, values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
        workbook.close()
        text = "\n".join(values)
        keywords = [
            "微信支付账单明细",
            "微信昵称",
            "交易类型",
            "交易单号",
        ]

        return sum(k in text for k in keywords) >= 3
    except Exception:
        return False

def load_alipay_bill(file_path: Path) -> list[dict]:
    """解析支付宝账单 CSV 文件。

    从文件中定位标题行，逐行读取并按照 ALIPAY_MAPPING 将原始字段映射为
    标准化的内部字段。每条记录会自动添加 ``source: "alipay"`` 标记。

    Args:
        file_path: 支付宝账单 CSV 文件路径，编码为 GBK。

    Returns:
        解析后的账单列表，每个元素为一个包含标准化字段的字典。

    Raises:
        ValueError: 文件中未找到以 "交易时间," 开头的标题行。
    """
    with open(file_path, "r", encoding="gbk", newline="") as f:
        lines = iter(f)
        headers = None
        for line in lines:
            line = line.strip()
            if line.startswith("交易时间,"):
                headers = next(csv.reader([line]))
                break
        if headers is None:
            raise ValueError("未找到支付宝账单标题")
        reader = csv.DictReader(lines, fieldnames=headers)
        result = []
        for row in reader:
            if not any(row.values()):
                continue
            item = {
                internal: row.get(source, "").strip()
                for source, internal in ALIPAY_MAPPING.items()
            }
            item["source"] = "alipay"
            result.append(item)
        return result

def load_wechat_bill(file_path: Path) -> list[dict]:
    """解析微信账单 Excel 文件。

    通过 openpyxl 读取 Excel 工作表，定位标题行后逐行解析，并按照
    WECHAT_MAPPING 将原始字段映射为标准化的内部字段。每条记录会自动添加
    ``source: "wechat"`` 标记。

    Args:
        file_path: 微信账单 Excel 文件路径（.xlsx 格式）。

    Returns:
        解析后的账单列表，每个元素为一个包含标准化字段的字典。

    Raises:
        ValueError: 工作表中未找到以 "交易时间" 作为首列的标题行。
    """
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = None
    for row in rows:
        values = [
            "" if cell is None else str(cell).strip()
            for cell in row
        ]
        if values and values[0] == "交易时间":
            headers = values
            break
    if headers is None:
        workbook.close()
        raise ValueError("未找到微信账单标题")
    result = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        raw = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = row[i] if i < len(row) else None
            raw[header] = "" if value is None else str(value).strip()
        item = {
            internal: raw.get(source, "")
            for source, internal in WECHAT_MAPPING.items()
        }
        item["source"] = "wechat"
        result.append(item)
    workbook.close()
    return result


# 检测器-解析器对列表，按优先级排列。
# 每个元素为 (detector, parser) 元组，detector 用于判断文件类型，
# parser 用于实际解析。遍历时匹配到第一个即停止。
PARSERS = [
    (is_alipay_bill, load_alipay_bill),
    (is_wechat_bill, load_wechat_bill),
]


def load_bills(directory: str) -> list[dict]:
    """加载并解析目录下的所有账单文件。

    遍历指定目录中的每个文件，使用注册的检测器自动识别文件类型
    （支付宝或微信），然后调用对应的解析器进行解析。无法识别的文件会被跳过。

    Args:
        directory: 存放账单文件的目录路径。

    Returns:
        所有账单文件解析结果的汇总列表，每个元素为一个包含标准化字段的字典。
    """
    bills = []
    for path in Path(directory).iterdir():
        if not path.is_file():
            continue
        matched = False
        for detector, parser in PARSERS:
            if detector(path):
                print(f"解析：{path.name}")
                bills.extend(parser(path))
                matched = True
                break
        if not matched:
            print(f"跳过：{path.name}")
    return bills


def sort_bills(bills: list[dict]) -> None:
    """按交易时间对账单列表进行原地升序排序。

    时间字段的格式为 ``%Y-%m-%d %H:%M:%S``。

    Args:
        bills: 待排序的账单列表，排序会直接修改原列表。
    """
    bills.sort(
        key=lambda bill: datetime.strptime(
            bill["time"],
            "%Y-%m-%d %H:%M:%S"
        )
    )


def export_json(bills: list[dict], output: str) -> None:
    """将账单列表导出为 JSON 文件。

    使用 UTF-8 编码，保留中文字符不转义，格式化缩进输出。

    Args:
        bills: 账单列表。
        output: 输出 JSON 文件的路径。
    """
    with open(output, "w", encoding="utf-8") as f:
        json.dump(
            bills,
            f,
            ensure_ascii=False,
            indent=2,
        )


if __name__ == "__main__":
    import sys
    import os

    if len(sys.argv) < 2:
        print("用法: python parse.py <账单目录> [输出路径]", file=sys.stderr)
        print("示例: python parse.py ~/Downloads/bills -o /tmp/bills.json", file=sys.stderr)
        sys.exit(1)

    input_dir = sys.argv[1]
    output_path = "/tmp/analyze_bill_data.json"

    # 解析 -o 参数指定输出路径
    args = sys.argv[2:]
    i = 0
    while i < len(args):
        if args[i] == "-o" and i + 1 < len(args):
            output_path = args[i + 1]
            i += 2
        else:
            i += 1

    if not os.path.isdir(input_dir):
        print(f"错误: 目录不存在 - {input_dir}", file=sys.stderr)
        sys.exit(1)

    # 删除旧数据，避免与新数据混淆
    if os.path.exists(output_path):
        os.remove(output_path)

    bills = load_bills(input_dir)

    if not bills:
        print("未识别到任何账单文件。请确保目录中包含支付宝 CSV 或微信 XLSX 账单文件。")
        sys.exit(0)

    sort_bills(bills)
    export_json(bills, output_path)

    # 统计信息
    alipay_count = sum(1 for b in bills if b["source"] == "alipay")
    wechat_count = sum(1 for b in bills if b["source"] == "wechat")
    time_min = bills[0]["time"]
    time_max = bills[-1]["time"]

    print(f"共导出 {len(bills)} 条账单记录 -> {output_path}")
    if alipay_count > 0:
        print(f"  支付宝: {alipay_count} 条")
    if wechat_count > 0:
        print(f"  微信: {wechat_count} 条")
    print(f"  时间范围: {time_min} ~ {time_max}")